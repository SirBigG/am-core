import logging
from datetime import datetime

import requests
from django.utils.translation import get_language
from openpyxl import load_workbook
from transliterate import slugify

from .models import Company, Country, Variety, VarietyCategory
from .parser_row_types import ActiveRegistryItem, InactiveRegistryItem

logger = logging.getLogger("django")

file_path = "media/derzhavnii-reiestr-sortiv-roslin-pridatnikh-dlia-poshirennia-v-ukrayini_19-03-2024.xlsx"


def parse_company(worksheet_number, api: bool = False, token: str = None, start_row: int = 2):
    # Make dict of countries
    countries = {country.short_slug: country.id for country in Country.objects.all()}
    workbook = load_workbook(filename=file_path)
    sheet = workbook.worksheets[worksheet_number]
    row_count = sheet.max_row
    counter = start_row - 1
    session = requests.Session()
    session.headers.update({"Authorization": f"Token {token}"})
    for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, values_only=True):
        if api:
            # send POST request to API
            response = session.post(
                "https://agromega.in.ua/api/registry/add-company/",
                json={"row": row},
                timeout=10,
            )
            if response.status_code != 200:
                logger.error(f"Error while sending data to API: {response.text}")
            counter += 1
            logger.info(f"Row {counter} of {row_count} sent to API")
            continue
        code = row[2]
        if Company.objects.filter(code=code).exists():
            continue
        company = Company.objects.create(
            name=row[3], original_name=row[4], code=code, country_id=countries.get(row[5].lower(), None)
        )
        company.save()


def parse_applicant(api: bool = False, token: str = None, start_row: int = 2):
    parse_company(2, api, token, start_row)


def parse_owner(api: bool = False, token: str = None, start_row: int = 2):
    parse_company(3, api, token, start_row)


def parse_breeder(api: bool = False, token: str = None, start_row: int = 2):
    parse_company(4, api, token, start_row)


def parse_varieties(api: bool = False, token: str = None, start_row: int = 2):
    countries = {country["short_slug"]: country["id"] for country in Country.objects.values("short_slug", "id")}
    companies = {company["code"]: company["id"] for company in Company.objects.values("code", "id")}
    base_categories = {
        category["title"]: category["id"] for category in VarietyCategory.objects.filter(level=1).values("title", "id")
    }
    children_categories = {
        category["title"]: category["id"] for category in VarietyCategory.objects.filter(level=2).values("title", "id")
    }
    workbook = load_workbook(filename=file_path)
    sheet = workbook.worksheets[0]
    row_count = sheet.max_row
    counter = start_row - 1
    session = requests.Session()
    session.headers.update({"Authorization": f"Token {token}"})
    for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, values_only=True):
        # if len of row is less than 37, then we have empty row and we need to set empty values
        if len(row) < 38:
            row = list(row) + ["" for _ in range(38 - len(row))]
        if len(row) > 38:
            row = row[:38]
        if api:
            # send POST request to API
            response = session.post(
                "https://agromega.in.ua/api/registry/add-active-variety/",
                json={"row": row},
                timeout=10,
            )
            if response.status_code != 200:
                logger.error(f"Error while sending data to API: {response.text}")
            counter += 1
            logger.info(f"Row {counter} of {row_count} sent to API")
            continue
        item = ActiveRegistryItem._make(row)
        base_category_title = item.base_category_title
        if base_category_title not in base_categories:
            # Check if category already exists
            slug = slugify(base_category_title, get_language())
            base_category = VarietyCategory.objects.filter(slug=slug).first()
            if base_category is None:
                base_category = VarietyCategory.objects.create(title=base_category_title)
                base_category.save()
            base_categories[base_category_title] = base_category.id
        base_category_id = base_categories[base_category_title]
        children_category_title = item.child_category_title
        if children_category_title not in children_categories:
            # Check if category already exists
            slug = slugify(children_category_title, get_language())
            children_category = VarietyCategory.objects.filter(slug=slug).first()
            if children_category is None:
                children_category = VarietyCategory.objects.create(
                    title=children_category_title, parent_id=base_category_id
                )
                children_category.save()
            children_categories[children_category_title] = children_category.id
        children_category_id = children_categories[children_category_title]
        title = item.title
        if Variety.objects.filter(title=title, category_id=children_category_id).exists():
            continue
        registration_country = item.registration_country
        if registration_country:
            registration_country = countries.get(item.registration_country.lower(), None)
        original_country = item.original_country
        if original_country:
            original_country = countries.get(item.original_country.lower(), None)
        variety = Variety.objects.create(
            title=title,
            title_original=item.title_original,
            application_number=item.application_number,
            registration_year=item.registration_year,
            recommended_zone=item.recommended_zone,
            direction_of_use=item.direction_of_use,
            ripeness_group=item.ripeness_group,
            quality=item.quality,
            registration_country_id=registration_country,
            original_country_id=original_country,
            applicant_id=companies.get(item.applicant, None),
            applicant2_id=companies.get(item.applicant2, None),
            owner_id=companies.get(item.owner, None),
            owner2_id=companies.get(item.owner2, None),
            breeder_id=companies.get(item.breeder, None),
            category_id=children_category_id,
        )
        variety.save()


def parse_inactive_varieties(api: bool = False, token: str = None, start_row: int = 2):
    countries = {country["short_slug"]: country["id"] for country in Country.objects.values("short_slug", "id")}
    companies = {company["code"]: company["id"] for company in Company.objects.values("code", "id")}
    base_categories = {
        category["title"]: category["id"] for category in VarietyCategory.objects.filter(level=1).values("title", "id")
    }
    children_categories = {
        category["title"]: category["id"] for category in VarietyCategory.objects.filter(level=2).values("title", "id")
    }
    workbook = load_workbook(filename=file_path)
    sheet = workbook.worksheets[1]
    row_count = sheet.max_row
    counter = start_row - 1
    session = requests.Session()
    session.headers.update({"Authorization": f"Token {token}"})
    for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, values_only=True):
        # if len of row is less than 37, then we have empty row and we need to set empty values
        if len(row) < 39:
            row = list(row) + ["" for _ in range(39 - len(row))]
        if len(row) > 39:
            row = row[:39]
        if api:
            # send POST request to API
            response = session.post(
                "https://agromega.in.ua/api/registry/add-inactive-variety/",
                json={"row": row},
                headers={"Authorization": f"Token {token}"},
                timeout=10,
            )
            if response.status_code != 200:
                logger.error(f"Error while sending data to API: {response.text}")
            counter += 1
            logger.info(f"Row {counter} of {row_count} sent to API")
            continue
        item = InactiveRegistryItem._make(row)
        end_date = item.end_date
        end_date_year = None
        if end_date:
            # parse end date by format dd.mm.yyyy
            end_date = datetime.strptime(end_date, "%d.%m.%Y").date()
            end_date_year = end_date.year
        base_category_title = item.base_category_title
        if base_category_title not in base_categories:
            # Check if category already exists
            slug = slugify(base_category_title, get_language())
            base_category = VarietyCategory.objects.filter(slug=slug).first()
            if base_category is None:
                base_category = VarietyCategory.objects.create(title=base_category_title)
                base_category.save()
            base_categories[base_category_title] = base_category.id
        base_category_id = base_categories[base_category_title]
        children_category_title = item.child_category_title
        if children_category_title not in children_categories:
            # Check if category already exists
            slug = slugify(children_category_title, get_language())
            children_category = VarietyCategory.objects.filter(slug=slug).first()
            if children_category is None:
                children_category = VarietyCategory.objects.create(
                    title=children_category_title, parent_id=base_category_id
                )
                children_category.save()
            children_categories[children_category_title] = children_category.id
        children_category_id = children_categories[children_category_title]
        title = item.title
        if Variety.objects.filter(title=title, category_id=children_category_id).exists():
            Variety.objects.filter(title=title, category_id=children_category_id).update(
                unregister_date=end_date, unregister_year=end_date_year, excluded=True
            )
            continue
        registration_country = item.registration_country
        if registration_country:
            registration_country = countries.get(item.registration_country.lower(), None)
        original_country = item.original_country
        if original_country:
            original_country = countries.get(item.original_country.lower(), None)
        variety = Variety.objects.create(
            title=title,
            title_original=item.title_original,
            application_number=item.application_number,
            registration_year=item.registration_year,
            recommended_zone=item.recommended_zone,
            direction_of_use=item.direction_of_use,
            ripeness_group=item.ripeness_group,
            quality=item.quality,
            registration_country_id=registration_country,
            original_country_id=original_country,
            applicant_id=companies.get(item.applicant, None),
            applicant2_id=companies.get(item.applicant2, None),
            owner_id=companies.get(item.owner, None),
            owner2_id=companies.get(item.owner2, None),
            breeder_id=companies.get(item.breeder, None),
            category_id=children_category_id,
            unregister_date=end_date,
            unregister_year=end_date_year,
            excluded=True,
        )
        variety.save()
